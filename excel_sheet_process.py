import streamlit as st
import pandas as pd
import io
import requests
import concurrent.futures
import re
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, Border, Side

# =====================================================================
# APP CONFIGURATION
# =====================================================================
st.set_page_config(page_title="Data Processing App", layout="wide")

st.title("Data Processing Application")
st.write("Please upload your PIM Master Data and SKU List to proceed.")

# Initialize session state to hold our processed file so it doesn't disappear on re-runs
if 'processed_data' not in st.session_state:
    st.session_state['processed_data'] = None
if 'file_name' not in st.session_state:
    st.session_state['file_name'] = ""

# =====================================================================
# FILE UPLOADERS
# =====================================================================
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. PIM Master Data")
    pim_file = st.file_uploader("Upload PIM Master Data (Excel)", type=["xls", "xlsx"], key="pim_upload")

with col2:
    st.subheader("2. SKU List")
    sku_file = st.file_uploader("Upload SKU List (Excel)", type=["xls", "xlsx"], key="sku_upload")

# Check if both files have been uploaded
if pim_file is not None and sku_file is not None:
    st.success("Files successfully uploaded!")
    st.divider()

    # User Selection for Processing Mode
    st.subheader("Select Processing Mode")
    process_mode = st.radio(
        "Which output would you like to generate?",
        ("Master File Data Processing (Assortment Formatted)", "Mainline & DF Data Processing")
    )

    # Process button prevents auto-running when the user is just toggling options
    if st.button("Run Processing"):
        with st.spinner("Processing data... this may take a moment."):
            try:
                # -------------------------------------------------------------
                # COMMON DATA READ & MERGE
                # -------------------------------------------------------------
                pim_df = pd.read_excel(pim_file)
                sku_df = pd.read_excel(sku_file)
                merged_df = pd.merge(sku_df[['SKU']], pim_df, on="SKU", how="left")
                num_rows = len(merged_df)

                # =============================================================
                # MODE 1: MASTER FILE DATA PROCESSING (ASSORTMENT)
                # =============================================================
                if process_mode == "Master File Data Processing (Assortment Formatted)":

                    def safe_get(col_name):
                        if col_name in merged_df.columns:
                            return merged_df[col_name].fillna("").astype(str)
                        return pd.Series([""] * num_rows)


                    if "UPC" in merged_df.columns:
                        merged_df["UPC"] = safe_get("UPC").str.replace(r'\.0$', '', regex=True)

                    columns = [
                        "SEASON", "Season\n(Remark Carryover)", "Picture", "Category",
                        "Style Color", "Group ", "Subgroup", "UPC", "Style Code",
                        "Style Description", "Color Code", "Color Name", "Size",
                        "Material Description", "", "USD", "RMB", "HKD", "TWD", "MOP",
                        "GC TTL Units", "Mainland TTL Units", "HK", "TW", "MC", "Buffer",
                        "Delivery", "RMB-Incl.VAT", "HKD.1", "TWD.1", "MOP.1", "Size.1"
                    ]

                    df_out = pd.DataFrame(index=range(num_rows), columns=columns).fillna("")

                    # Static & Dynamic Values
                    df_out["SEASON"] = "F26"
                    df_out["Season\n(Remark Carryover)"] = "F26"
                    df_out["Style Color"] = safe_get("SKU")
                    df_out["Style Code"] = safe_get("SKU")
                    df_out["UPC"] = safe_get("UPC")
                    df_out["Category"] = safe_get("Product Type")
                    df_out["Picture"] = safe_get("Main Image")

                    # Conditional Mappings
                    is_jewelry = safe_get("Product Type").str.upper().str.contains("JEWELRY", na=False)

                    df_out["Group "] = safe_get("Platform")
                    if "Group" in merged_df.columns:
                        df_out.loc[is_jewelry, "Group "] = merged_df.loc[is_jewelry, "Group"].fillna("").astype(str)

                    df_out["Subgroup"] = safe_get("Platform")
                    if "Primary Color Jewelry" in merged_df.columns:
                        df_out.loc[is_jewelry, "Subgroup"] = merged_df.loc[is_jewelry, "Primary Color Jewelry"].fillna(
                            "").astype(str)

                    platform_str = safe_get("Platform")
                    size_str = safe_get("Case Size").str.replace(r'\.0$', '', regex=True)
                    df_out["Style Description"] = (platform_str + " " + size_str).str.strip()

                    if "Product Name" in merged_df.columns:
                        def clean_jewelry_description(row):
                            p_name = str(row.get("Product Name", "")) if pd.notna(row.get("Product Name")) else ""
                            brand = str(row.get("Brand", "")) if pd.notna(row.get("Brand")) else ""
                            if p_name and brand:
                                cleaned = re.sub(re.escape(brand), '', p_name, flags=re.IGNORECASE)
                                return cleaned.strip()
                            return p_name.strip()


                        merged_df["Cleaned Product Name"] = merged_df.apply(clean_jewelry_description, axis=1)
                        df_out.loc[is_jewelry, "Style Description"] = merged_df.loc[
                            is_jewelry, "Cleaned Product Name"].fillna("").astype(str)
                    else:
                        df_out.loc[is_jewelry, "Style Description"] = ""

                    df_out["Color Code"] = "-"
                    if "SAP Color" in merged_df.columns:
                        df_out.loc[is_jewelry, "Color Code"] = merged_df.loc[is_jewelry, "SAP Color"].fillna("").astype(
                            str)

                    df_out["Color Name"] = safe_get("Ecomm Top Ring Color")
                    if "Silhouette Jewelry" in merged_df.columns:
                        df_out.loc[is_jewelry, "Color Name"] = merged_df.loc[is_jewelry, "Silhouette Jewelry"].fillna(
                            "").astype(str)

                    df_out["Size"] = safe_get("Case Size")
                    df_out["Size.1"] = safe_get("Case Size")
                    df_out.loc[is_jewelry, "Size"] = "-"

                    df_out["Material Description"] = safe_get("Ecomm Case Material")
                    if "Primary Material Jewelry" in merged_df.columns:
                        df_out.loc[is_jewelry, "Material Description"] = merged_df.loc[
                            is_jewelry, "Primary Material Jewelry"].fillna("").astype(str)

                    df_out = df_out.fillna("").astype(str)


                    # High-Speed Image Downloader
                    def fetch_image(url):
                        try:
                            response = requests.get(str(url).strip(), timeout=5)
                            if response.status_code == 200:
                                return url, response.content
                        except:
                            pass
                        return url, None


                    unique_urls = [url for url in df_out["Picture"].unique() if
                                   url and str(url).strip().startswith("http")]
                    image_cache = {}

                    if unique_urls:
                        st.info("Downloading images...")
                        progress_bar = st.progress(0)
                        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
                            futures = {executor.submit(fetch_image, url): url for url in unique_urls}
                            for i, future in enumerate(concurrent.futures.as_completed(futures)):
                                url = futures[future]
                                image_cache[url] = future.result()[1]
                                progress_bar.progress((i + 1) / len(unique_urls))
                        progress_bar.empty()

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_out.to_excel(writer, index=False, startrow=5, sheet_name="F26 LICENSEE")
                        worksheet = writer.sheets["F26 LICENSEE"]

                        header_font = Font(bold=True, color="000000")
                        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        data_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        data_left = Alignment(horizontal='left', vertical='center', wrap_text=False)
                        thin_border = Border(
                            left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
                        )

                        worksheet.cell(row=2, column=16, value="PRICE")
                        worksheet.cell(row=2, column=21, value="ORDER UNITS")
                        worksheet.cell(row=2, column=27, value="DELIVERY")
                        worksheet.cell(row=2, column=28, value="COST")
                        worksheet.cell(row=3, column=1, value="F26")
                        worksheet.cell(row=5, column=21, value=2400)
                        worksheet.cell(row=5, column=22, value=1545)
                        worksheet.cell(row=5, column=23, value=164)
                        worksheet.cell(row=5, column=24, value=549)
                        worksheet.cell(row=5, column=25, value=142)

                        for col_letter in [chr(i) for i in range(65, 91)] + ['AA', 'AB', 'AC', 'AD', 'AE', 'AF']:
                            worksheet.column_dimensions[col_letter].width = 15
                        worksheet.column_dimensions['C'].width = 10.5
                        worksheet.column_dimensions['O'].width = 3

                        for col_num in range(1, 33):
                            cell = worksheet.cell(row=6, column=col_num)
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border

                        for col_num in [16, 21, 27, 28]:
                            cell = worksheet.cell(row=2, column=col_num)
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border

                        for col_num in range(21, 26):
                            cell = worksheet.cell(row=5, column=col_num)
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border

                        for idx, url in enumerate(df_out["Picture"]):
                            row_number = idx + 7
                            worksheet.row_dimensions[row_number].height = 60
                            for col_num in range(1, 33):
                                cell = worksheet.cell(row=row_number, column=col_num)
                                cell.alignment = data_center if col_num not in [10, 14] else data_left
                                cell.border = thin_border

                            cell_ref = f"C{row_number}"
                            worksheet[cell_ref].value = ""
                            img_bytes = image_cache.get(url)
                            if img_bytes:
                                try:
                                    img = OpenpyxlImage(io.BytesIO(img_bytes))
                                    img.width = 65
                                    img.height = 65
                                    worksheet.add_image(img, cell_ref)
                                except Exception:
                                    worksheet[cell_ref].value = "Img Format Error"
                            elif pd.notna(url) and str(url).strip().startswith("http"):
                                worksheet[cell_ref].value = "Link Error"

                    # Save to Session State
                    st.session_state['processed_data'] = output.getvalue()
                    st.session_state['file_name'] = "Assortment_Formatted_Output.xlsx"

                # =============================================================
                # MODE 2: MAINLINE & DF DATA PROCESSING
                # =============================================================
                elif process_mode == "Mainline & DF Data Processing":

                    if "UPC" in merged_df.columns:
                        merged_df["UPC"] = merged_df["UPC"].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)

                    # 1. Item Master Sheet
                    df_item_master = pd.DataFrame(index=range(num_rows), columns=[
                        "Purpose Code", "Total Number of line items", "Assigned Identification",
                        "Style Number", "UPC number", "Color Code", "Size code",
                        "Style Description/Platform Name", "Style Color", "Style Size",
                        "Product Description Code"
                    ])
                    df_item_master["Purpose Code"] = 2
                    df_item_master["Total Number of line items"] = 1
                    df_item_master["Assigned Identification"] = 1
                    df_item_master["Style Number"] = merged_df["SKU"]
                    df_item_master["UPC number"] = merged_df["UPC"]
                    df_item_master["Color Code"] = merged_df["SAP Color"]
                    df_item_master["Size code"] = merged_df["Case Size"]
                    df_item_master["Style Description/Platform Name"] = merged_df["Platform"]

                    # Check if 'Product Type' indicates Jewelry, and if so, map to 'Group' instead
                    if "Product Type" in merged_df.columns and "Group" in merged_df.columns:
                        is_jewelry = merged_df["Product Type"].fillna("").astype(str).str.upper().str.contains(
                            "JEWELRY", na=False)
                        df_item_master.loc[is_jewelry, "Style Description/Platform Name"] = merged_df.loc[
                            is_jewelry, "Group"]
                    # Default mapping to 'Case Color'
                    df_item_master["Style Color"] = merged_df["Case Color"]

                    # Check if 'Product Type' indicates Jewelry, and if so, map to 'Primary Color Jewelry'
                    if "Product Type" in merged_df.columns and "Primary Color Jewelry" in merged_df.columns:
                        is_jewelry = merged_df["Product Type"].fillna("").astype(str).str.upper().str.contains(
                            "JEWELRY", na=False)
                        df_item_master.loc[is_jewelry, "Style Color"] = merged_df.loc[
                            is_jewelry, "Primary Color Jewelry"]
                    df_item_master["Style Size"] = merged_df["Case Size"]

                    # 2. Cost- ML Sheet
                    df_cost_ml = pd.DataFrame(index=range(num_rows), columns=[
                        "Relation", "AccountCode", "AccountRelation", "ItemCode",
                        "UPC", "SKU", "From date", "To date", "Unit", "Amount", "Currency"
                    ])
                    df_cost_ml["Relation"] = "000::Price (Purch)"
                    df_cost_ml["AccountCode"] = "002::All"
                    df_cost_ml["ItemCode"] = "000::Table"
                    df_cost_ml["Unit"] = "Pcs"
                    df_cost_ml["Currency"] = "CNY"
                    df_cost_ml["SKU"] = merged_df["SKU"]
                    df_cost_ml["UPC"] = merged_df["UPC"]

                    # 3. PO ML-FP Sheet
                    df_po_ml = pd.DataFrame(index=range(num_rows), columns=[
                        "vendor code", "UPC", "Style", "Color", "Size"
                    ])
                    df_po_ml["vendor code"] = "V21100001"
                    df_po_ml["Style"] = merged_df["SKU"]
                    df_po_ml["Color"] = merged_df["Case Color"]
                    df_po_ml["Size"] = merged_df["Case Size"]
                    df_po_ml["UPC"] = merged_df["UPC"]

                    # 4. Price-ML Sheet
                    df_price_ml = pd.DataFrame(index=range(num_rows), columns=[
                        "Relation", "AccountCode", "AccountRelation", "ItemCode",
                        "UPC", "SKU", "From date", "To date", "Unit", "Amount", "Currency"
                    ])
                    df_price_ml["Relation"] = "004::Price (sales)"
                    df_price_ml["AccountCode"] = "002::All"
                    df_price_ml["ItemCode"] = "000::Table"
                    df_price_ml["Unit"] = "Pcs"
                    df_price_ml["Currency"] = "CNY"
                    df_price_ml["SKU"] = merged_df["SKU"]
                    df_price_ml["UPC"] = merged_df["UPC"]

                    # Package into Excel
                    sheets_to_export = {
                        "Item Master": df_item_master,
                        "Cost- ML": df_cost_ml,
                        "PO ML-FP": df_po_ml,
                        "Price-ML": df_price_ml
                    }

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for sheet_name, df in sheets_to_export.items():
                            df.to_excel(writer, index=False, sheet_name=sheet_name)

                    # Save to Session State
                    st.session_state['processed_data'] = output.getvalue()
                    st.session_state['file_name'] = "MK_Mainline_Output.xlsx"

            except KeyError as e:
                st.error(
                    f"Mapping Error: Could not find the expected column in your uploaded files. Please ensure {e} exists.")
            except Exception as e:
                st.error(f"An error occurred: {e}")

# =====================================================================
# DOWNLOAD SECTION
# =====================================================================
# Only show the download button if we have successfully saved data in session state
if st.session_state['processed_data'] is not None:
    st.divider()
    st.subheader("Download Output")
    st.success(f"Your processed Excel file ({st.session_state['file_name']}) is ready!")

    st.download_button(
        label="📥 Download Processed Excel File",
        data=st.session_state['processed_data'],
        file_name=st.session_state['file_name'],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_btn_{st.session_state['file_name']}"  # <--- THIS IS THE FIX
    )

elif pim_file is None or sku_file is None:
    st.info("Waiting for both files to be uploaded...")